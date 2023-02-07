from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

wb = load_workbook("borclar.xlsx")
ws = wb["Sheet1"]

ws.title = "Borclar"